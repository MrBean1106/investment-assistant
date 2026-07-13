"""Excel import/export helpers built on openpyxl.

`columns` is an ordered list of (field, header) tuples. Export writes the
header row and one row per dict; import reads by header name back into dicts
(JSON list fields like tags arrive as comma-joined strings and are split by
the caller).
"""

from io import BytesIO
from typing import Optional

import openpyxl
from openpyxl import Workbook


def rows_to_xlsx(rows: list[dict], columns: list[tuple[str, str]]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    ws.append([c[1] for c in columns])
    for row in rows:
        ws.append([row.get(c[0], "") for c in columns])
    # Auto-width for readability
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 12), 40)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def xlsx_to_rows(content: bytes, columns: list[tuple[str, str]]) -> list[dict]:
    wb = openpyxl.load_workbook(BytesIO(content), read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []
    header_map = {str(h).strip(): field for field, h in columns if h}
    result: list[dict] = []
    for raw in rows_iter:
        if raw is None:
            continue
        d: dict = {}
        for i, val in enumerate(raw):
            if i >= len(header):
                break
            hname = str(header[i]).strip()
            field = header_map.get(hname)
            if field:
                d[field] = "" if val is None else val
        if any(v not in (None, "") for v in d.values()):
            result.append(d)
    return result


XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
